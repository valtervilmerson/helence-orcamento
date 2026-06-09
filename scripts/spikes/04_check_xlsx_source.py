"""Spike: o PDF foi gerado de um Excel (metadata Title = TABELA DE PRECO 01-2025_REUNIOES.xlsx).
Ha um data/prices.xlsx na pasta -- vamos checar se e a mesma fonte (estrutura de abas/colunas),
o que mudaria MUITO a estrategia de extracao (xlsx estruturado vs. PDF de layout livre)."""
import openpyxl

PATH = r"C:\projects\codex\helence\helence-orcamento\data\prices.xlsx"
wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
print("Abas:", wb.sheetnames)

for name in wb.sheetnames[:3]:
    ws = wb[name]
    print("=" * 80)
    print(f"Aba: {name} | dimensoes: {ws.dimensions} | max_row={ws.max_row} max_col={ws.max_column}")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True)):
        print(f"  linha {i+1}: {row}")
