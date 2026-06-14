from __future__ import annotations

import json
import re
from collections import Counter
from datetime import datetime, UTC
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
SOURCE_PATH = next(DATA_DIR.glob("*SOLUCOES ACUSTICAS.xlsx"))
OUTPUT_PATH = Path(__file__).resolve().parent / "importacao_solucoes_acusticas.json"

FAMILY = "Soluções Acústicas"
PLACEHOLDER_FINISH = "Sem acabamento aplicável"
PLACEHOLDER_FINISH_NOTE = (
    "A planilha não explicita acabamento para este item; o parser preencheu "
    "o placeholder 'Sem acabamento aplicável'."
)

WOODISH_FINISHES = {
    "Amendoa": "madeirado",
    "Carvalho": "madeirado",
    "Gianduia": "madeirado",
    "Itapua": "madeirado",
    "Nogueira Cadiz": "madeirado",
}


def norm_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ").strip())


def sku_str(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, int):
        text = str(value)
        return text if 6 <= len(text) <= 12 else None
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
        return text if 6 <= len(text) <= 12 else None
    if isinstance(value, str):
        text = value.strip()
        if re.fullmatch(r"\d{6,12}(?:\.0+)?", text):
            return text.split(".")[0]
    return None


def is_price(value: object) -> bool:
    return isinstance(value, (int, float)) and 0 < float(value) < 100000


def normalize_dimension(text: str | None) -> str | None:
    if not text:
        return None

    value = norm_text(text).upper()
    if not value:
        return None

    if re.fullmatch(r"D\s*\d+", value):
        number = re.sub(r"\D", "", value)
        return f"D{number}"

    normalized = value.replace("MM", "").replace(" ", "").replace("X", "x")
    normalized = re.sub(r"(?<=\d)[HLP](?=[x\d])", "", normalized)
    normalized = re.sub(r"[^0-9xD]", "", normalized)

    if re.fullmatch(r"D\d+", normalized):
        return normalized
    if re.fullmatch(r"\d+(x\d+){1,2}", normalized):
        return normalized
    return None


def title_pt(text: str) -> str:
    return text.strip().title()


def dedupe_notes(*parts: str | None) -> str | None:
    values = []
    seen = set()
    for part in parts:
        if not part:
            continue
        key = part.strip()
        if not key or key in seen:
            continue
        seen.add(key)
        values.append(key)
    if not values:
        return None
    return " ".join(values)


def strip_dimension_from_description(description: str, dimension: str | None) -> str:
    text = description
    if dimension:
        text = re.sub(re.escape(dimension), "", text, flags=re.IGNORECASE).strip(" -")
    text = re.sub(r"\b\d+(?:x\d+){1,2}\b", "", text, flags=re.IGNORECASE).strip(" -")
    return re.sub(r"\s+", " ", text).strip()


def derive_product_context(
    sheet_name: str,
    current_context: str,
    description: str,
    dimension: str | None,
) -> str:
    generic_contexts = {
        "BAFFLE",
        "DIVISOR DE AMBIENTE",
        "MOSAICO",
        "MOSAICO 3D",
        "NUVEM",
        "PAINEL FRISADO",
        "PAINEL FRISADO MOLDURA",
    }

    if current_context and current_context.upper() not in generic_contexts:
        return title_pt(current_context)

    cleaned = re.sub(
        r"^(KIT (?:PARA |DE )?MONTAGEM DE?\s+)",
        "",
        description,
        flags=re.IGNORECASE,
    )
    cleaned = strip_dimension_from_description(cleaned, dimension)

    if cleaned:
        return title_pt(cleaned)
    if dimension:
        return title_pt(f"{sheet_name} {dimension}")
    return title_pt(sheet_name)


def infer_component_type(sheet_name: str, description: str) -> str:
    desc = description.lower()
    sheet = sheet_name.lower()
    if "kit" in desc:
        return "Kit de Montagem"
    if "estrutura" in desc:
        return "Estrutura"
    if "placa acústica" in desc:
        return "Placa Acústica"
    if "painel divisor" in desc:
        return "Painel Divisor"
    if "ripado" in desc or "ripado" in sheet:
        return "Painel Ripado"
    if "frisado" in desc or "frisado" in sheet:
        return "Painel Frisado"
    if "baffle" in desc or "baffle" in sheet:
        return "Baffle"
    if "divisor de ambiente" in desc or "divisor de ambiente" in sheet:
        return "Divisor de Ambiente"
    if "mosaico" in desc or "mosaico" in sheet:
        return "Mosaico"
    if "nuvem" in desc or "nuvem" in sheet:
        return "Nuvem"
    if "biombo" in desc or "biombo" in sheet:
        return "Biombo"
    return "Solução Acústica"


def finish_group_for(finish: str) -> str:
    return WOODISH_FINISHES.get(finish, "outro")


def build_item(
    *,
    ref: str,
    sheet_name: str,
    current_context: str,
    description: str,
    dimension: str | None,
    finish: str,
    sku: str,
    price: float,
    confidence: float,
    notes: str | None,
) -> dict[str, object]:
    return {
        "ref": ref,
        "family": FAMILY,
        "product_context": derive_product_context(sheet_name, current_context, description, dimension),
        "component_type": infer_component_type(sheet_name, description),
        "description": title_pt(description),
        "dimension": dimension or "NA",
        "finish": finish,
        "finish_group": finish_group_for(finish),
        "sku": sku,
        "price": round(float(price), 2),
        "currency": "BRL",
        "confidence": confidence,
        "notes": notes,
    }


def parse_grid_group(
    *,
    rows: list[list[object]],
    sheet_name: str,
    allowed_cols: set[int],
) -> list[dict[str, object]]:
    items: list[dict[str, object]] = []
    nonempty_idx = [
        idx
        for idx, row in enumerate(rows)
        if any(col < len(row) and norm_text(row[col]) for col in allowed_cols)
    ]
    current_context = sheet_name
    last_dim_by_col: dict[int, str] = {}

    for pos, idx in enumerate(nonempty_idx):
        row = rows[idx]
        visible_texts = [norm_text(row[col]) for col in sorted(allowed_cols) if col < len(row) and norm_text(row[col])]
        first_col = min(allowed_cols)
        first_text = norm_text(row[first_col]) if first_col < len(row) else ""

        if (
            len(visible_texts) == 1
            and first_text
            and "IMAGENS" not in first_text.upper()
            and "DESCRIÇÃO TÉCNICA" not in first_text.upper()
            and first_text.upper() != "TOTAL"
            and not sku_str(first_text)
        ):
            current_context = first_text

        for col in allowed_cols:
            if col >= len(row):
                continue
            dimension = normalize_dimension(norm_text(row[col]))
            if dimension:
                last_dim_by_col[col] = dimension

        sku_cols = [col for col in allowed_cols if col < len(row) and sku_str(row[col])]
        if not sku_cols:
            continue

        next_nonempty_idx = None
        for future_idx in nonempty_idx[pos + 1 : pos + 5]:
            if any(col < len(rows[future_idx]) and rows[future_idx][col] is not None for col in sku_cols):
                next_nonempty_idx = future_idx
                break

        if next_nonempty_idx is None:
            continue
        if any(
            col < len(rows[next_nonempty_idx]) and sku_str(rows[next_nonempty_idx][col])
            for col in sku_cols
        ):
            continue
        if not any(
            col < len(rows[next_nonempty_idx]) and is_price(rows[next_nonempty_idx][col])
            for col in sku_cols
        ):
            continue

        for col in sku_cols:
            sku = sku_str(row[col])
            price = rows[next_nonempty_idx][col] if col < len(rows[next_nonempty_idx]) else None
            if not sku or not is_price(price):
                continue

            description = None
            dimension = last_dim_by_col.get(col)
            for previous_idx in reversed(nonempty_idx[max(0, pos - 8) : pos]):
                if col >= len(rows[previous_idx]):
                    continue
                candidate = norm_text(rows[previous_idx][col])
                if not candidate or candidate.upper().startswith("TOTAL"):
                    continue
                if not description and not normalize_dimension(candidate):
                    description = candidate
                    break

            if not description:
                continue
            if not dimension:
                dimension = normalize_dimension(description)

            items.append(
                build_item(
                    ref=f"{sheet_name}!R{idx + 1}C{col + 1}",
                    sheet_name=sheet_name,
                    current_context=current_context,
                    description=description,
                    dimension=dimension,
                    finish=PLACEHOLDER_FINISH,
                    sku=sku,
                    price=float(price),
                    confidence=0.9,
                    notes=PLACEHOLDER_FINISH_NOTE,
                )
            )

    return items


def parse_panel_divisor(rows: list[list[object]]) -> list[dict[str, object]]:
    items: list[dict[str, object]] = []
    current_context = "Painel Divisor Acústico"

    for idx, row in enumerate(rows):
        first = norm_text(row[0]) if len(row) > 0 else ""
        second = norm_text(row[1]) if len(row) > 1 else ""
        third = row[2] if len(row) > 2 else None

        if (
            first
            and "PAINEL DIVISOR" in first.upper()
            and "MODELO" not in first.upper()
            and "PLATAFORMA" not in first.upper()
            and "IMAGENS" not in first.upper()
        ):
            current_context = first

        sku = sku_str(third)
        if not sku or idx + 2 >= len(rows):
            continue

        dimension = normalize_dimension(norm_text(rows[idx + 1][1]) if len(rows[idx + 1]) > 1 else "")
        support = norm_text(rows[idx + 2][1]) if len(rows[idx + 2]) > 1 else ""
        price = rows[idx + 2][2] if len(rows[idx + 2]) > 2 else None
        if not dimension or not is_price(price):
            continue

        description = second or current_context
        if support:
            description = f"{description} - {support}"

        items.append(
            build_item(
                ref=f"PAINEL DIVISOR!R{idx + 1}C3",
                sheet_name="PAINEL DIVISOR",
                current_context=current_context,
                description=description,
                dimension=dimension,
                finish=PLACEHOLDER_FINISH,
                sku=sku,
                price=float(price),
                confidence=0.94,
                notes=PLACEHOLDER_FINISH_NOTE,
            )
        )

    return items


def parse_painel_ripado(rows: list[list[object]]) -> list[dict[str, object]]:
    items: list[dict[str, object]] = []
    finishes = [title_pt(norm_text(value)) for value in rows[2][2:11]]
    nonempty_idx = [
        idx
        for idx, row in enumerate(rows)
        if any(norm_text(value) for value in row[:11])
    ]

    for pos, idx in enumerate(nonempty_idx):
        row = rows[idx]
        if not any(col < len(row) and sku_str(row[col]) for col in range(2, 11)):
            continue

        description = norm_text(row[1]) if len(row) > 1 else ""
        if not description:
            continue

        if pos + 2 >= len(nonempty_idx):
            continue
        dimension_row = rows[nonempty_idx[pos + 1]]
        price_row = rows[nonempty_idx[pos + 2]]
        dimension = normalize_dimension(norm_text(dimension_row[1]) if len(dimension_row) > 1 else "")
        if not dimension:
            continue

        for col, finish in zip(range(2, 11), finishes, strict=False):
            if col >= len(row) or col >= len(price_row):
                continue

            sku = sku_str(row[col])
            price = price_row[col]
            if not sku or not is_price(price):
                continue

            items.append(
                build_item(
                    ref=f"PAINEL RIPADO!R{idx + 1}C{col + 1}",
                    sheet_name="PAINEL RIPADO",
                    current_context="Painel Acústico Ripado",
                    description=description,
                    dimension=dimension,
                    finish=finish,
                    sku=sku,
                    price=float(price),
                    confidence=0.98,
                    notes=None,
                )
            )

    return items


def parse_biombo(rows: list[list[object]]) -> list[dict[str, object]]:
    items: list[dict[str, object]] = []

    structure_sku = sku_str(rows[5][2])
    panel_sku = sku_str(rows[5][4])
    dimension = normalize_dimension(norm_text(rows[7][1]))
    structure_price = rows[8][10]
    panel_price = rows[8][12]
    total_price = rows[8][13]

    if structure_sku and dimension and is_price(structure_price):
        items.append(
            build_item(
                ref="BIOMBO!R6C3",
                sheet_name="BIOMBO",
                current_context="Biombo Acústico",
                description="Estrutura para Biombo",
                dimension=dimension,
                finish="Preto",
                sku=structure_sku,
                price=float(structure_price),
                confidence=0.92,
                notes=None,
            )
        )

    if panel_sku and dimension and is_price(panel_price):
        items.append(
            build_item(
                ref="BIOMBO!R6C5",
                sheet_name="BIOMBO",
                current_context="Biombo Acústico",
                description="Placa Acústica para Biombo",
                dimension=dimension,
                finish=PLACEHOLDER_FINISH,
                sku=panel_sku,
                price=float(panel_price),
                confidence=0.88,
                notes=PLACEHOLDER_FINISH_NOTE,
            )
        )

    if structure_sku and panel_sku and dimension and is_price(total_price):
        items.append(
            build_item(
                ref="BIOMBO!R9C14",
                sheet_name="BIOMBO",
                current_context="Biombo Acústico",
                description="Biombo Acústico Completo",
                dimension=dimension,
                finish="Preto",
                sku=f"{structure_sku}+{panel_sku}",
                price=float(total_price),
                confidence=0.72,
                notes=(
                    "SKU composto pelo parser a partir de estrutura + placa; "
                    "a planilha informa apenas o total do conjunto."
                ),
            )
        )

    return items


def dedupe_exact_items(items: list[dict[str, object]]) -> list[dict[str, object]]:
    deduped: dict[tuple[object, ...], dict[str, object]] = {}
    for item in items:
        key = (
            item["product_context"],
            item["component_type"],
            item["description"],
            item["dimension"],
            item["finish"],
            item["sku"],
            item["price"],
        )
        deduped.setdefault(key, item)
    return list(deduped.values())


def build_contract() -> dict[str, object]:
    workbook = load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    items: list[dict[str, object]] = []

    grid_specs = {
        "BAFFLE": set(range(0, 10)),
        "DIVISOR DE AMBIENTE": set(range(0, 10)),
        "MOSAICO": set(range(0, 10)),
        "MOSAICO 3D": set(range(0, 10)),
        "NUVEM": set(range(0, 10)),
    }

    for sheet_name, allowed_cols in grid_specs.items():
        rows = [list(row) for row in workbook[sheet_name].iter_rows(values_only=True)]
        items.extend(parse_grid_group(rows=rows, sheet_name=sheet_name, allowed_cols=allowed_cols))

    frisado_groups = {
        "PAINEL FRISADO MOLDURA": [set([0, 1]), set([11, 12])],
        "PAINEL FRISADO": [set([0, 1]), set([11, 12])],
    }
    for sheet_name, groups in frisado_groups.items():
        rows = [list(row) for row in workbook[sheet_name].iter_rows(values_only=True)]
        for allowed_cols in groups:
            items.extend(parse_grid_group(rows=rows, sheet_name=sheet_name, allowed_cols=allowed_cols))

    items.extend(parse_panel_divisor([list(row) for row in workbook["PAINEL DIVISOR"].iter_rows(values_only=True)]))
    items.extend(parse_painel_ripado([list(row) for row in workbook["PAINEL RIPADO"].iter_rows(values_only=True)]))
    items.extend(parse_biombo([list(row) for row in workbook["BIOMBO"].iter_rows(values_only=True)]))

    final_items = dedupe_exact_items(items)
    final_items.sort(key=lambda item: (str(item["product_context"]), str(item["component_type"]), str(item["sku"])))

    return {
        "contract_version": "1.0",
        "source": {
            "description": SOURCE_PATH.name,
            "generated_by": "codex-parser-manual",
            "generated_at": datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        },
        "items": final_items,
    }


def main() -> None:
    contract = build_contract()
    OUTPUT_PATH.write_text(
        json.dumps(contract, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    counts = Counter(item["product_context"] for item in contract["items"])
    print(f"Arquivo gerado: {OUTPUT_PATH}")
    print(f"Itens: {len(contract['items'])}")
    print("Primeiros contextos:")
    for context, total in counts.most_common(10):
        print(f"- {context}: {total}")


if __name__ == "__main__":
    main()
