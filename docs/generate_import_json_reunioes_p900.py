"""Piloto: gera o JSON de importação para a aba P900 da tabela de Reuniões.

Escopo deste piloto: apenas a aba ``P900`` (mesas de reunião 1200x900 a
5400x900, tampos + estruturas). Após validação do mapeamento de campos pelo
time, o parser será generalizado para as demais abas (P1000-P1800 e
acessórios).
"""

from __future__ import annotations

import json
import re
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
SOURCE_PATH = DATA_DIR / "TABELA DE PRECO 01-2026_REUNIOES.xlsx"
OUTPUT_PATH = Path(__file__).resolve().parent / "importacao_reunioes_p900_piloto.json"

SHEET_NAME = "P900"
LAST_ROW = 3624  # última linha não vazia da aba (1-indexada)

FAMILY = "Mesas de Reunião"
KNOWN_PRODUCT_CONTEXT = "Reunião 1200x900"

# Colunas 4-12 (0-indexadas) = acabamentos do cabeçalho da aba (linha 3).
FINISH_BY_COL = {
    4: "Argila",
    5: "Branco",
    6: "Preto",
    7: "Gianduia",
    8: "Amendoa",
    9: "Carvalho",
    10: "Nogueira Cadiz",
    11: "Grafite",
    12: "Itapua",
}


def norm_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ").strip())


def sku_str(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, int):
        text = str(value)
        return text if 6 <= len(text) <= 12 else None
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
        return text if 6 <= len(text) <= 12 else None
    if isinstance(value, str):
        text = value.strip()
        if re.fullmatch(r"\d{6,12}(?:\.0+)?", text):
            return text.split(".")[0]
    return None


def is_price(value: object) -> bool:
    return isinstance(value, (int, float)) and 0 < float(value) < 100000


def dedupe_notes(*parts: str | None) -> str | None:
    values = []
    seen = set()
    for part in parts:
        if not part:
            continue
        key = part.strip()
        if not key or key in seen:
            continue
        seen.add(key)
        values.append(key)
    if not values:
        return None
    return " ".join(values)


def expand_interm(text: str) -> str:
    return re.sub(r"\b[Ii]nterm\.?\b", "Intermediário", text)


def build_item(
    *,
    ref: str,
    product_context: str,
    component_type: str,
    description: str,
    dimension: str,
    finish: str,
    finish_group: str | None,
    sku: str,
    price: float,
    confidence: float,
    notes: str | None,
) -> dict[str, object]:
    item: dict[str, object] = {
        "ref": ref,
        "family": FAMILY,
        "product_context": product_context,
        "component_type": component_type,
        "description": description,
        "dimension": dimension,
        "finish": finish,
        "sku": sku,
        "price": round(float(price), 2),
        "currency": "BRL",
        "confidence": confidence,
        "notes": notes,
    }
    if finish_group:
        item["finish_group"] = finish_group
    return item


def parse_p900(rows: list[list[object]]) -> list[dict[str, object]]:
    items: list[dict[str, object]] = []
    current_width: str | None = None
    current_section = "TAMPOS"
    active_header: dict[str, object] | None = None

    for i, row in enumerate(rows):
        col0 = norm_text(row[0]) if len(row) > 0 else ""
        col3 = norm_text(row[3]) if len(row) > 3 else ""

        width_match = re.match(r"Reuni[aã]o (\d+x900)", col0)
        if width_match:
            current_width = width_match.group(1)
            current_section = "TAMPOS"
            active_header = None
            # A própria linha do marcador de largura também carrega o
            # primeiro grupo de SKUs de tampo (FOSCO/Simples) — não usar
            # `continue` aqui, para não perder esse grupo.

        if col0.upper() == "ESTRUTURAS":
            current_section = "ESTRUTURAS"
            active_header = None
            continue

        if current_width is None:
            continue

        if current_section == "TAMPOS":
            if (
                col3.startswith("Tampo ")
                and i + 3 < len(rows)
                and all(sku_str(row[c]) for c in range(4, 13))
            ):
                price_row = rows[i + 3]
                if not all(is_price(price_row[c]) for c in range(4, 13)):
                    continue

                caixa_label = norm_text(price_row[3])
                is_known_product = current_width == "1200x900"
                base_confidence = 0.96 if is_known_product else 0.93
                product_note = (
                    None
                    if is_known_product
                    else (
                        f"Produto 'Reunião {current_width}' ainda não existe no "
                        "catálogo (linha de produto nova)."
                    )
                )

                for col in range(4, 13):
                    items.append(
                        build_item(
                            ref=f"L{i + 1},{i + 4}",
                            product_context=f"Reunião {current_width}",
                            component_type="Tampo",
                            description=(
                                f"{col3} Para Estrutura Reunião {current_width} {caixa_label}"
                            ),
                            dimension=current_width,
                            finish=FINISH_BY_COL[col],
                            finish_group=None,
                            sku=sku_str(row[col]) or "",
                            price=float(price_row[col]),
                            confidence=base_confidence,
                            notes=product_note,
                        )
                    )
            continue

        # current_section == "ESTRUTURAS"
        if col3 in ("REUNIÃO 5050", "APOIO CREDENZA 5050"):
            if all(row[c] is not None for c in (4, 5, 6)):
                active_header = {
                    "idx": i,
                    "finish_names": [norm_text(row[4]), norm_text(row[5]), norm_text(row[6])],
                }
            else:
                active_header = None
            continue

        if not col3:
            continue

        sku_cols = [c for c in range(4, 13) if sku_str(row[c]) is not None]
        if len(sku_cols) not in (3, 9):
            continue
        if sku_cols[:3] != [4, 5, 6]:
            continue
        if i + 3 >= len(rows):
            continue

        price_row = rows[i + 3]
        if not norm_text(price_row[3]):
            continue
        if not all(is_price(price_row[c]) for c in sku_cols):
            continue

        desc_extra1 = norm_text(rows[i + 1][3]) if i + 1 < len(rows) else ""
        desc_extra2 = norm_text(rows[i + 2][3]) if i + 2 < len(rows) else ""

        description = expand_interm(
            re.sub(
                r"\s+",
                " ",
                f"{col3} {desc_extra1} {desc_extra2} {current_width}".strip(),
            )
        )

        is_credenza = "Apoio Credenza" in col3
        component_type = "Estrutura Apoio Credenza" if is_credenza else "Estrutura"

        is_known_product = current_width == "1200x900"
        product_note = (
            None
            if is_known_product
            else (
                f"Produto 'Reunião {current_width}' ainda não existe no "
                "catálogo (linha de produto nova)."
            )
        )

        if is_credenza:
            ct_note = "Tipo de componente 'Estrutura Apoio Credenza' é novo no catálogo."
        else:
            ct_note = (
                "Tipo de componente 'Estrutura' ainda não está associado a "
                "produtos da família 'Mesas de Reunião' no catálogo "
                "(apenas 'Tampo')."
            )

        if len(sku_cols) == 3:
            finish_names = ["Prata", "Preto", "Branco"]
            if active_header is not None and active_header["idx"] == i - 1:
                ref = f"L{int(active_header['idx']) + 1},{i + 1},{i + 4}"
            else:
                ref = f"L{i + 1},{i + 4}"
            confidence = 0.80 if is_credenza else 0.85
        else:
            finish_names = [FINISH_BY_COL[c] for c in sku_cols]
            ref = f"L{i + 1},{i + 4}"
            confidence = 0.85

        ambiguous_note = None
        if len(sku_cols) == 9 and re.search(r"PRATA|PRETO|BRANCO", desc_extra1.upper()):
            ambiguous_note = (
                "Estrutura com 9 variações de acabamento (mesmas colunas dos "
                f"tampos) sob o rótulo '{desc_extra1}'; confirmar mapeamento "
                "de acabamento com o time."
            )

        for col, finish in zip(sku_cols, finish_names, strict=True):
            finish_group = "metalico" if finish == "Prata" else None
            finish_note = (
                "Acabamento 'Prata' não consta no catálogo; sugerido grupo 'metalico'."
                if finish == "Prata"
                else None
            )

            items.append(
                build_item(
                    ref=ref,
                    product_context=f"Reunião {current_width}",
                    component_type=component_type,
                    description=description,
                    dimension=current_width,
                    finish=finish,
                    finish_group=finish_group,
                    sku=sku_str(row[col]) or "",
                    price=float(price_row[col]),
                    confidence=confidence,
                    notes=dedupe_notes(ct_note, finish_note, ambiguous_note, product_note),
                )
            )

    return items


def build_contract() -> dict[str, object]:
    workbook = load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    sheet = workbook[SHEET_NAME]
    rows = [list(row) for row in sheet.iter_rows(max_row=LAST_ROW, values_only=True)]

    items = parse_p900(rows)
    items.sort(
        key=lambda item: (
            str(item["product_context"]),
            str(item["component_type"]),
            str(item["sku"]),
        )
    )

    return {
        "contract_version": "1.0",
        "source": {
            "description": f"{SOURCE_PATH.name} ({SHEET_NAME}, piloto)",
            "generated_by": "codex-parser-manual",
            "generated_at": (
                datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")
            ),
        },
        "items": items,
    }


def main() -> None:
    contract = build_contract()
    OUTPUT_PATH.write_text(
        json.dumps(contract, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    items = contract["items"]
    counts = Counter(item["component_type"] for item in items)
    print(f"Arquivo gerado: {OUTPUT_PATH}")
    print(f"Itens: {len(items)}")
    print("Por tipo de componente:")
    for component_type, total in counts.most_common():
        print(f"- {component_type}: {total}")


if __name__ == "__main__":
    main()
