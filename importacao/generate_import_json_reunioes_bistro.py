"""Gera o JSON de importação da planilha Reuniões Bistrô.

TABELA DE PRECO 01-2026_REUNIOES BISTRO.xlsx — 7 abas produtivas (as demais
são de apoio e são ignoradas: CUSTOS_SISTEMA, ÍNDICE REVENDA, COEF, PERC_VENDA).

Layout comum a todas as abas (blocos de 4 linhas, começando na linha 6):
  Linha 0: col1=descrição + cols finish=SKUs
  Linha 1: col1=dimensão (ex.: '1200x500x1000mm' ou '1200x500X25mm')
  Linha 2: col1=rótulo do pé (ex.: 'Pé Painel') — vazio para tampos
  Linha 3: cols finish=preços

Aba 'Reunião Lounge Bistrô' tem variação: a descrição se distribui por 2
linhas antes da dimensão; os preços ficam junto ao rótulo do pé (linha 3).

Acabamentos:
  - Tampos e Estruturas pé painel/alum → 9 madeirados (cols 2-10):
    Argila, Branco, Preto, Gianduia, Itapua, Amendoa, Carvalho,
    Nogueira Cadiz, Grafite
  - Estrutura pé aço 5050 → 4 metálicos (cols 2-5):
    Prata, Preto, Branco, Grafite
"""
from __future__ import annotations

import json
import re
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
SOURCE_PATH = DATA_DIR / "TABELA DE PRECO 01-2026_REUNIOES BISTRO.xlsx"
OUTPUT_PATH = Path(__file__).resolve().parent / "importacao_reunioes_bistro.json"

FAMILY = "Mesas de Reunião"

# Acabamentos madeirados — tampos e estruturas pé painel/alum (cols 2-10)
FINISH_MADEIRADO: dict[int, str] = {
    2: "Argila",
    3: "Branco",
    4: "Preto",
    5: "Gianduia",
    6: "Itapua",
    7: "Amendoa",
    8: "Carvalho",
    9: "Nogueira Cadiz",
    10: "Grafite",
}

# Acabamentos metálicos — estrutura pé aço 5050 (cols 2-5)
FINISH_METALICO: dict[int, str] = {
    2: "Prata",
    3: "Preto",
    4: "Branco",
    5: "Grafite",
}

# Abas produtivas: nome da aba → configuração de parsing
SHEET_CONFIG: list[dict] = [
    {
        "name": "Tampos Reunião Bistrô",
        "component_type": "Tampo",
        "finish_by_col": FINISH_MADEIRADO,
        "finish_group": "madeirado",
        "kind": "standard",
    },
    {
        "name": "Estrutura Reu Bistrô Pe Painel",
        "component_type": "Estrutura",
        "finish_by_col": FINISH_MADEIRADO,
        "finish_group": "madeirado",
        "kind": "standard",
    },
    {
        "name": "Estrut Reu Bistrô Pe Alum Fosco",
        "component_type": "Estrutura",
        "finish_by_col": FINISH_MADEIRADO,
        "finish_group": "madeirado",
        "kind": "standard",
    },
    {
        "name": "Estrut Reu Bistrô Pe Alum Preto",
        "component_type": "Estrutura",
        "finish_by_col": FINISH_MADEIRADO,
        "finish_group": "madeirado",
        "kind": "standard",
    },
    {
        "name": "Estrut Reu Bistrô Pe Alum Branc",
        "component_type": "Estrutura",
        "finish_by_col": FINISH_MADEIRADO,
        "finish_group": "madeirado",
        "kind": "standard",
    },
    {
        "name": "Estr Reu Bistrô Pe Aço 5050",
        "component_type": "Estrutura",
        "finish_by_col": FINISH_METALICO,
        "finish_group": "metalico",
        "kind": "standard",
    },
    {
        "name": "Reunião Lounge Bistrô",
        "component_type": "Reunião Lounge Bistrô",
        "finish_by_col": FINISH_MADEIRADO,
        "finish_group": "madeirado",
        "kind": "lounge",
    },
]


# ---------------------------------------------------------------------------
# Utilitários
# ---------------------------------------------------------------------------

def norm_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ").strip())


def sku_str(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, int):
        text = str(value)
        return text if 6 <= len(text) <= 12 else None
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
        return text if 6 <= len(text) <= 12 else None
    if isinstance(value, str):
        text = value.strip()
        if re.fullmatch(r"\d{6,12}(?:\.0+)?", text):
            return text.split(".")[0]
    return None


def is_price(value: object) -> bool:
    return isinstance(value, (int, float)) and 0 < float(value) < 1_000_000


def norm_dimension(raw: str) -> str:
    """Normaliza string de dimensão da planilha.

    '1200x500x1000mm' → '1200x500x1000'
    '1200x500X25mm'   → '1200x500x25'
    'Diam 700mm'      → '700MM'
    """
    s = raw.strip()
    m = re.match(r"Diam\s*(\d+)\s*mm", s, re.IGNORECASE)
    if m:
        return f"{m.group(1)}MM"
    m = re.match(r"(\d+)[xX](\d+)[xX](\d+)\s*(?:mm)?", s, re.IGNORECASE)
    if m:
        return f"{m.group(1)}x{m.group(2)}x{m.group(3)}"
    m = re.match(r"(\d+)[xX](\d+)\s*(?:mm)?", s, re.IGNORECASE)
    if m:
        return f"{m.group(1)}x{m.group(2)}"
    return s


def extract_lxp(raw: str) -> str | None:
    """Extrai 'LxP' de strings como '1200x500x1000mm' ou '1200x500X25mm'."""
    m = re.match(r"(\d+)[xX](\d+)", raw.strip())
    if m:
        return f"{m.group(1)}x{m.group(2)}"
    return None


def is_dimension_str(text: str) -> bool:
    return bool(
        re.match(r"\d+[xX]\d+", text)
        or re.match(r"Diam\s*\d+", text, re.IGNORECASE)
    )


def build_item(
    *,
    ref: str,
    product_context: str | None,
    component_type: str,
    description: str,
    dimension: str,
    finish: str,
    finish_group: str,
    sku: str | None,
    price: float,
    confidence: float,
    notes: str | None,
) -> dict[str, object]:
    item: dict[str, object] = {
        "ref": ref,
        "family": FAMILY,
        "product_context": product_context,
        "component_type": component_type,
        "description": description,
        "dimension": dimension,
        "finish": finish,
        "finish_group": finish_group,
        "sku": sku,
        "price": round(float(price), 2),
        "currency": "BRL",
        "confidence": confidence,
        "notes": notes,
    }
    return item


# ---------------------------------------------------------------------------
# Parsers
# ---------------------------------------------------------------------------

def _get(row: list[object], col: int) -> object:
    return row[col] if col < len(row) else None


def parse_standard_sheet(
    rows: list[list[object]],
    sheet_name: str,
    *,
    component_type: str,
    finish_by_col: dict[int, str],
    finish_group: str,
) -> list[dict[str, object]]:
    """Analisa abas com blocos de 4 linhas (tampos e estruturas)."""
    items: list[dict[str, object]] = []
    finish_cols = sorted(finish_by_col.keys())
    i = 5  # dados começam na linha 6 (índice 5)

    while i + 3 < len(rows):
        row0, row1, row2, row3 = rows[i], rows[i + 1], rows[i + 2], rows[i + 3]

        # --- detecção de bloco válido ---
        # row0: deve ter SKUs nas colunas de acabamento
        skus = {c: sku_str(_get(row0, c)) for c in finish_cols}
        if not any(skus.values()):
            i += 1
            continue

        # row1: deve ter string de dimensão em col1
        dim_raw = norm_text(_get(row1, 1))
        if not is_dimension_str(dim_raw):
            i += 1
            continue

        # row3: deve ter preços nas colunas de acabamento
        prices = {c: _get(row3, c) for c in finish_cols}
        if not any(is_price(prices[c]) for c in finish_cols):
            i += 1
            continue

        # --- extração ---
        desc_main = norm_text(_get(row0, 1))
        pe_label = norm_text(_get(row2, 1))  # vazio para tampos
        dimension = norm_dimension(dim_raw)

        lxp = extract_lxp(dim_raw)
        product_context = f"Reunião Bistrô {lxp}" if lxp else None

        description = " ".join(p for p in [desc_main, pe_label] if p)

        notes = (
            "Linha de produto 'Reunião Bistrô' é nova no catálogo."
        )
        confidence = 0.85

        for col in finish_cols:
            sku = skus[col]
            price = prices[col]
            if not is_price(price):
                continue
            items.append(
                build_item(
                    ref=f"BISTRO.xlsx!{sheet_name}!L{i + 1},{i + 4}",
                    product_context=product_context,
                    component_type=component_type,
                    description=description,
                    dimension=dimension,
                    finish=finish_by_col[col],
                    finish_group=finish_group,
                    sku=sku,
                    price=float(price),
                    confidence=confidence,
                    notes=notes,
                )
            )
        i += 4

    return items


def parse_lounge_sheet(
    rows: list[list[object]],
    sheet_name: str,
) -> list[dict[str, object]]:
    """Analisa a aba 'Reunião Lounge Bistrô'.

    Cada bloco de 4 linhas:
      linha 0: col1='Reunião Redonda ' + SKUs cols 2-10
      linha 1: col1='Lounge Bistrô '
      linha 2: col1='Diam 700mm'
      linha 3: col1='Pé Alum POLIDO' (ou 'Pé Aço PRATA' etc.) + preços cols 2-10
    """
    items: list[dict[str, object]] = []
    finish_cols = sorted(FINISH_MADEIRADO.keys())
    i = 5

    while i + 3 < len(rows):
        row0, row1, row2, row3 = rows[i], rows[i + 1], rows[i + 2], rows[i + 3]

        # SKUs em row0 cols 2-10
        skus = {c: sku_str(_get(row0, c)) for c in finish_cols}
        if not any(skus.values()):
            i += 1
            continue

        # Dimensão em row2 col1
        dim_raw = norm_text(_get(row2, 1))
        if not is_dimension_str(dim_raw):
            i += 1
            continue

        # Preços em row3 cols 2-10
        prices = {c: _get(row3, c) for c in finish_cols}
        if not any(is_price(prices[c]) for c in finish_cols):
            i += 1
            continue

        desc_part1 = norm_text(_get(row0, 1))
        desc_part2 = norm_text(_get(row1, 1))
        pe_label = norm_text(_get(row3, 1))  # ex.: "Pé Alum POLIDO" ou "Pé Aço PRATA"
        dimension = norm_dimension(dim_raw)

        description = " ".join(p for p in [desc_part1, desc_part2, pe_label] if p)

        notes = (
            "Tipo de componente 'Reunião Lounge Bistrô' é novo no catálogo. "
            "Linha de produto 'Reunião Lounge Bistrô' é nova no catálogo."
        )
        confidence = 0.80

        for col in finish_cols:
            sku = skus[col]
            price = prices[col]
            if not is_price(price):
                continue
            items.append(
                build_item(
                    ref=f"BISTRO.xlsx!{sheet_name}!L{i + 1},{i + 4}",
                    product_context="Reunião Lounge Bistrô",
                    component_type="Reunião Lounge Bistrô",
                    description=description,
                    dimension=dimension,
                    finish=FINISH_MADEIRADO[col],
                    finish_group="madeirado",
                    sku=sku,
                    price=float(price),
                    confidence=confidence,
                    notes=notes,
                )
            )
        i += 4

    return items


# ---------------------------------------------------------------------------
# Geração do contrato
# ---------------------------------------------------------------------------

def build_contract() -> dict[str, object]:
    wb = load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    all_items: list[dict[str, object]] = []

    for cfg in SHEET_CONFIG:
        sheet_name = cfg["name"]
        ws = wb[sheet_name]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]

        if cfg["kind"] == "lounge":
            items = parse_lounge_sheet(rows, sheet_name)
        else:
            items = parse_standard_sheet(
                rows,
                sheet_name,
                component_type=cfg["component_type"],
                finish_by_col=cfg["finish_by_col"],
                finish_group=cfg["finish_group"],
            )

        display_name = re.sub(r"[\x80-\xff]", "?", sheet_name)
        print(f"  {display_name}: {len(items)} itens")
        all_items.extend(items)

    all_items.sort(
        key=lambda item: (
            str(item.get("product_context") or ""),
            str(item["component_type"]),
            str(item["dimension"]),
            str(item["finish"]),
        )
    )

    return {
        "contract_version": "1.0",
        "source": {
            "description": SOURCE_PATH.name,
            "generated_by": "codex-parser-manual",
            "generated_at": (
                datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")
            ),
        },
        "items": all_items,
    }


def main() -> None:
    print("Processando abas:")
    contract = build_contract()
    OUTPUT_PATH.write_text(
        json.dumps(contract, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    items = contract["items"]
    counts = Counter(str(item["component_type"]) for item in items)
    print(f"\nArquivo gerado: {OUTPUT_PATH}")
    print(f"Itens: {len(items)}")
    print("Por tipo de componente:")
    for ct, total in counts.most_common():
        print(f"  {ct}: {total}")

    contexts = Counter(str(item.get("product_context") or "(avulso)") for item in items)
    print(f"\nProdutos distintos: {len(contexts)}")
    for ctx, total in sorted(contexts.items()):
        print(f"  {ctx}: {total} variações")


if __name__ == "__main__":
    main()
