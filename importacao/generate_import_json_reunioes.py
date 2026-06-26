"""Gera o JSON de importação completo da planilha de Reuniões.

Generaliza o piloto (``generate_import_json_reunioes_p900.py``, validado
para a aba ``P900``) para as 10 abas de mesas retangulares/quadradas da
planilha ``TABELA DE PRECO 01-2026_REUNIOES.xlsx``: ``P900`` a ``P1800``
(profundidade 900mm a 1800mm), cada uma cobrindo larguras de 1200 a
5400mm em passos de 100mm — 43 larguras por aba, mesmo layout em todas
(cabeçalho de acabamentos na linha 3, blocos de 4 linhas
nome+códigos/observação/observação/preços).

Fora do escopo deste gerador (estrutura de planilha bastante diferente,
ver `docs/10-contrato-importacao-json.md` seção 4 e levantamento manual):
a aba ``REUNIÕES GERAL_ACESSÓRIOS`` (mesas redondas, conectores STAFF) e
as abas de apoio ``CUSTOS_SISTEMA``/``ÍNDICE REVENDA``/``COEF``/
``PERC_VENDA`` (ignoradas, não representam itens vendáveis).
"""

from __future__ import annotations

import json
import re
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
SOURCE_PATH = DATA_DIR / "TABELA DE PRECO 01-2026_REUNIOES.xlsx"
OUTPUT_PATH = Path(__file__).resolve().parent / "importacao_reunioes.json"

SHEET_NAMES = [
    "P900", "P1000", "P1100", "P1200", "P1300",
    "P1400", "P1500", "P1600", "P1700", "P1800",
]

FAMILY = "Mesas de Reunião"
KNOWN_PRODUCT_CONTEXT = "Reunião 1200x900"

# Colunas 4-12 (0-indexadas) = acabamentos do cabeçalho (linha 3), igual
# em todas as 10 abas (confirmado por amostragem).
FINISH_BY_COL = {
    4: "Argila",
    5: "Branco",
    6: "Preto",
    7: "Gianduia",
    8: "Amendoa",
    9: "Carvalho",
    10: "Nogueira Cadiz",
    11: "Grafite",
    12: "Itapua",
}


def norm_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ").strip())


def sku_str(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, int):
        text = str(value)
        return text if 6 <= len(text) <= 12 else None
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
        return text if 6 <= len(text) <= 12 else None
    if isinstance(value, str):
        text = value.strip()
        if re.fullmatch(r"\d{6,12}(?:\.0+)?", text):
            return text.split(".")[0]
    return None


def is_price(value: object) -> bool:
    return isinstance(value, (int, float)) and 0 < float(value) < 100000


def dedupe_notes(*parts: str | None) -> str | None:
    values = []
    seen = set()
    for part in parts:
        if not part:
            continue
        key = part.strip()
        if not key or key in seen:
            continue
        seen.add(key)
        values.append(key)
    if not values:
        return None
    return " ".join(values)


def expand_interm(text: str) -> str:
    return re.sub(r"\b[Ii]nterm\.?\b", "Intermediário", text)


def build_item(
    *,
    ref: str,
    product_context: str,
    component_type: str,
    description: str,
    dimension: str,
    finish: str,
    finish_group: str | None,
    sku: str,
    price: float,
    confidence: float,
    notes: str | None,
) -> dict[str, object]:
    item: dict[str, object] = {
        "ref": ref,
        "family": FAMILY,
        "product_context": product_context,
        "component_type": component_type,
        "description": description,
        "dimension": dimension,
        "finish": finish,
        "sku": sku,
        "price": round(float(price), 2),
        "currency": "BRL",
        "confidence": confidence,
        "notes": notes,
    }
    if finish_group:
        item["finish_group"] = finish_group
    return item


def parse_sheet(rows: list[list[object]], sheet_name: str) -> list[dict[str, object]]:
    items: list[dict[str, object]] = []
    current_width: str | None = None
    current_section = "TAMPOS"
    active_header: dict[str, object] | None = None

    for i, row in enumerate(rows):
        col0 = norm_text(row[0]) if len(row) > 0 else ""
        col3 = norm_text(row[3]) if len(row) > 3 else ""

        width_match = re.match(r"Reuni[aã]o (\d+x\d+)", col0)
        if width_match:
            current_width = width_match.group(1)
            current_section = "TAMPOS"
            active_header = None
            # A própria linha do marcador de largura também carrega o
            # primeiro grupo de SKUs de tampo (FOSCO/Simples) — não usar
            # `continue` aqui, para não perder esse grupo.

        if col0.upper() == "ESTRUTURAS":
            current_section = "ESTRUTURAS"
            active_header = None
            continue

        if current_width is None:
            continue

        if current_section == "TAMPOS":
            col2 = norm_text(row[2]) if len(row) > 2 else ""
            # Anomalia de layout localizada (aba P1400, larguras 1400x1400 e
            # 1500x1400): o rótulo "Tampo ..." aparece uma coluna à esquerda
            # do padrão (col2 em vez de col3) — as colunas de SKU/preço
            # (4-12) não são afetadas. Detectar e marcar para revisão em vez
            # de descartar silenciosamente o bloco.
            label_col: int | None
            if col3.startswith("Tampo "):
                label_col = 3
            elif col2.startswith("Tampo "):
                label_col = 2
            else:
                label_col = None

            if (
                label_col is not None
                and i + 3 < len(rows)
                and all(sku_str(row[c]) for c in range(4, 13))
            ):
                price_row = rows[i + 3]
                if not all(is_price(price_row[c]) for c in range(4, 13)):
                    continue

                tampo_label = norm_text(row[label_col])
                # O rótulo da "caixa" (linha de preço) tem coluna própria —
                # o desvio de layout do rótulo "Tampo ..." (linha de
                # cabeçalho) não implica que a linha de preço também
                # desviou. Detectar de forma independente em vez de
                # reaproveitar `label_col`, ou a descrição sai incompleta
                # quando só uma das duas linhas está deslocada (caso real:
                # aba P1400, largura 1400x1400).
                price_col3 = norm_text(price_row[3]) if len(price_row) > 3 else ""
                price_col2 = norm_text(price_row[2]) if len(price_row) > 2 else ""
                caixa_label = price_col3 or price_col2
                product_context = f"Reunião {current_width}"
                is_known_product = product_context == KNOWN_PRODUCT_CONTEXT
                layout_note = (
                    "Layout de coluna desviado no bloco de tampo (rótulo na "
                    "coluna anterior ao padrão); confirmar mapeamento antes "
                    "de publicar."
                    if label_col != 3
                    else None
                )
                base_confidence = (
                    0.6
                    if layout_note
                    else (0.96 if is_known_product else 0.93)
                )
                product_note = (
                    None
                    if is_known_product
                    else (
                        f"Produto '{product_context}' ainda não existe no "
                        "catálogo (linha de produto nova)."
                    )
                )

                for col in range(4, 13):
                    items.append(
                        build_item(
                            ref=f"REUNIOES.xlsx!{sheet_name}!L{i + 1},{i + 4}",
                            product_context=product_context,
                            component_type="Tampo",
                            description=(
                                f"{tampo_label} Para Estrutura Reunião "
                                f"{current_width} {caixa_label}"
                            ),
                            dimension=current_width,
                            finish=FINISH_BY_COL[col],
                            finish_group="madeirado",
                            sku=sku_str(row[col]) or "",
                            price=float(price_row[col]),
                            confidence=base_confidence,
                            notes=dedupe_notes(layout_note, product_note),
                        )
                    )
            continue

        # current_section == "ESTRUTURAS"
        if col3 in ("REUNIÃO 5050", "APOIO CREDENZA 5050"):
            if all(row[c] is not None for c in (4, 5, 6)):
                active_header = {
                    "idx": i,
                    "finish_names": [norm_text(row[4]), norm_text(row[5]), norm_text(row[6])],
                }
            else:
                active_header = None
            continue

        if not col3:
            continue

        sku_cols = [c for c in range(4, 13) if sku_str(row[c]) is not None]
        if len(sku_cols) not in (3, 9):
            continue
        if sku_cols[:3] != [4, 5, 6]:
            continue
        if i + 3 >= len(rows):
            continue

        price_row = rows[i + 3]
        if not norm_text(price_row[3]):
            continue
        if not all(is_price(price_row[c]) for c in sku_cols):
            continue

        desc_extra1 = norm_text(rows[i + 1][3]) if i + 1 < len(rows) else ""
        desc_extra2 = norm_text(rows[i + 2][3]) if i + 2 < len(rows) else ""

        description = expand_interm(
            re.sub(
                r"\s+",
                " ",
                f"{col3} {desc_extra1} {desc_extra2} {current_width}".strip(),
            )
        )

        is_credenza = "Apoio Credenza" in col3
        component_type = "Estrutura Apoio Credenza" if is_credenza else "Estrutura"

        product_context = f"Reunião {current_width}"
        is_known_product = product_context == KNOWN_PRODUCT_CONTEXT
        product_note = (
            None
            if is_known_product
            else (
                f"Produto '{product_context}' ainda não existe no "
                "catálogo (linha de produto nova)."
            )
        )

        if is_credenza:
            ct_note = "Tipo de componente 'Estrutura Apoio Credenza' é novo no catálogo."
        else:
            ct_note = (
                "Tipo de componente 'Estrutura' ainda não está associado a "
                "produtos da família 'Mesas de Reunião' no catálogo "
                "(apenas 'Tampo')."
            )

        if len(sku_cols) == 3:
            finish_names = ["Prata", "Preto", "Branco"]
            if active_header is not None and active_header["idx"] == i - 1:
                ref = f"REUNIOES.xlsx!{sheet_name}!L{int(active_header['idx']) + 1},{i + 1},{i + 4}"
            else:
                ref = f"REUNIOES.xlsx!{sheet_name}!L{i + 1},{i + 4}"
            confidence = 0.80 if is_credenza else 0.85
        else:
            finish_names = [FINISH_BY_COL[c] for c in sku_cols]
            ref = f"REUNIOES.xlsx!{sheet_name}!L{i + 1},{i + 4}"
            confidence = 0.85

        ambiguous_note = None
        if len(sku_cols) == 9 and re.search(r"PRATA|PRETO|BRANCO", desc_extra1.upper()):
            ambiguous_note = (
                "Estrutura com 9 variações de acabamento (mesmas colunas dos "
                f"tampos) sob o rótulo '{desc_extra1}'; confirmar mapeamento "
                "de acabamento com o time."
            )

        for col, finish in zip(sku_cols, finish_names, strict=True):
            # Confirmado com o time: Estrutura pé aço/Apoio Credenza (3 cores)
            # é metálica; Estrutura pé painel/alumínio/direto (9 cores) usa os
            # mesmos acabamentos madeira do tampo.
            finish_group = "metalico" if len(sku_cols) == 3 else "madeirado"
            finish_note = None

            items.append(
                build_item(
                    ref=ref,
                    product_context=product_context,
                    component_type=component_type,
                    description=description,
                    dimension=current_width,
                    finish=finish,
                    finish_group=finish_group,
                    sku=sku_str(row[col]) or "",
                    price=float(price_row[col]),
                    confidence=confidence,
                    notes=dedupe_notes(ct_note, finish_note, ambiguous_note, product_note),
                )
            )

    return items


def build_contract() -> dict[str, object]:
    workbook = load_workbook(SOURCE_PATH, read_only=True, data_only=True)

    all_items: list[dict[str, object]] = []
    for sheet_name in SHEET_NAMES:
        sheet = workbook[sheet_name]
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        items = parse_sheet(rows, sheet_name)
        print(f"  {sheet_name}: {len(items)} itens")
        all_items.extend(items)

    all_items.sort(
        key=lambda item: (
            str(item["dimension"]),
            str(item["product_context"]),
            str(item["component_type"]),
            str(item["sku"]),
        )
    )

    return {
        "contract_version": "1.0",
        "source": {
            "description": (
                f"{SOURCE_PATH.name} (abas {SHEET_NAMES[0]}-{SHEET_NAMES[-1]})"
            ),
            "generated_by": "codex-parser-manual",
            "generated_at": (
                datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")
            ),
        },
        "items": all_items,
    }


def main() -> None:
    print("Processando abas:")
    contract = build_contract()
    OUTPUT_PATH.write_text(
        json.dumps(contract, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    items = contract["items"]
    counts = Counter(item["component_type"] for item in items)
    print(f"\nArquivo gerado: {OUTPUT_PATH}")
    print(f"Itens: {len(items)}")
    print("Por tipo de componente:")
    for component_type, total in counts.most_common():
        print(f"- {component_type}: {total}")


if __name__ == "__main__":
    main()
