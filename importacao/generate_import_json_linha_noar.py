"""Gera o JSON de importação da planilha TABELA DE PRECO 01-2026_LINHA NOAR.xlsx.

Layout uniforme em todas as abas (exceto SOFÁ NEO):
  - Linha 1: título da aba
  - Linha 2: "*imagens meramente ilustrativas."
  - Linha 3: cabeçalho — MODELO | DESCRIÇÃO | Faixa1 | Faixa2 | Faixa3
  - Linhas 4+: blocos de 4 linhas por item:
      (i)   col1 = nome do produto (col0 = rótulo de grupo/MODELO quando muda)
      (i+1) col1 = dimensão "L x P x H mm"
      (i+2) col1 = detalhe de estrutura ("Pé Aço com madeira …")
      (i+3) col1 = detalhe de configuração ("com Pé Intermediário", "com gav …")
             col2 = preço faixa 1, col3 = preço faixa 2, col4 = preço faixa 3

Os produtos da Linha NOAR são vendidos como unidades completas (tampo + estrutura
integrados), ao contrário da Linha Reuniões onde tampo e estrutura têm SKUs
separados. Por isso:
  - `sku` = null em todos os itens
  - `component_type` = categoria do produto (ex. "Mesa Orgânica", "Plataforma")
  - `product_context` = null (produto autônomo, não sub-componente)
  - `family` = "Linha Noar"

Faixas de acabamento: a planilha usa preços por grupo de cor (ex. "Argila / Branco"
= um único preço para ambos os acabamentos). Os nomes das faixas são mantidos como
`finish` no JSON — são entidades novas no catálogo, que cairão na fila de revisão
para o time decidir como mapear para acabamentos individuais.

SOFÁ NEO (aba especial): layout diferente, cols 5/6 com erros `#REF!`; apenas a
coluna de preço válida (col2 = LINHO / COURO ECO) é extraída. 3 itens gerados.
"""

from __future__ import annotations

import json
import re
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
SOURCE_PATH = DATA_DIR / "TABELA DE PRECO 01-2026_LINHA NOAR.xlsx"
OUTPUT_PATH = Path(__file__).resolve().parent / "importacao_linha_noar.json"

FAMILY = "Linha Noar"

# Configuração por aba: component_type, faixas de acabamento (col_idx → nome)
# e finish_group correspondente.
SHEET_CONFIGS: dict[str, dict] = {
    "MESA ORGÂNICA": {
        "component_type": "Mesa Orgânica",
        "finishes": {
            2: ("Argila / Branco",          "madeirado"),
            3: ("Mescla de Cores",           "madeirado"),
            4: ("Madeirado",                 "madeirado"),
        },
    },
    "ESTAÇÃO DE TRABALHO": {
        "component_type": "Estação de Trabalho",
        "finishes": {
            2: ("Argila / Branco",           "madeirado"),
            3: ("Mescla de Cores",           "madeirado"),
            4: ("Madeirado",                 "madeirado"),
        },
    },
    "CREDENZA NOAR": {
        "component_type": "Credenza",
        "finishes": {
            2: ("Argila / Branco",           "madeirado"),
            3: ("Mescla de Cores",           "madeirado"),
            4: ("Madeirado",                 "madeirado"),
        },
    },
    "REUNIÃO ORGÂNICA": {
        "component_type": "Mesa de Reunião Orgânica",
        "finishes": {
            2: ("Argila / Branco",           "madeirado"),
            3: ("Mescla de Cores",           "madeirado"),
            4: ("Madeirado",                 "madeirado"),
        },
    },
    "PLATAFORMA CANTO RETO 1200 PROF": {
        "component_type": "Plataforma",
        "finishes": {
            2: ("Argila / Branco",           "madeirado"),
            3: ("Gianduia / Grafite / Preto","madeirado"),
            4: ("Madeirado",                 "madeirado"),
        },
    },
    "PLATAFORMA CANTO RETO 1400 PROF": {
        "component_type": "Plataforma",
        "finishes": {
            2: ("Argila / Branco",           "madeirado"),
            3: ("Gianduia / Grafite / Preto","madeirado"),
            4: ("Madeirado",                 "madeirado"),
        },
    },
    "PLATAFORMA CANTO CURVO 1200 P": {
        "component_type": "Plataforma",
        "finishes": {
            2: ("Argila / Branco",           "madeirado"),
            3: ("Gianduia / Grafite / Preto","madeirado"),
            4: ("Madeirado",                 "madeirado"),
        },
    },
    "PLATAFORMA CANTO CURVO 1400 P": {
        "component_type": "Plataforma",
        "finishes": {
            2: ("Argila / Branco",           "madeirado"),
            3: ("Gianduia / Grafite / Preto","madeirado"),
            4: ("Madeirado",                 "madeirado"),
        },
    },
    "REUNIÃO RETANGULAR CANTO RETO": {
        "component_type": "Mesa de Reunião Retangular",
        "finishes": {
            2: ("Argila / Branco",           "madeirado"),
            3: ("Gianduia / Grafite / Preto","madeirado"),
            4: ("Madeirado",                 "madeirado"),
        },
    },
    "REUNIÃO RETANGULAR CANTO CURVO": {
        "component_type": "Mesa de Reunião Retangular",
        "finishes": {
            2: ("Argila / Branco",           "madeirado"),
            3: ("Gianduia / Grafite / Preto","madeirado"),
            4: ("Madeirado",                 "madeirado"),
        },
    },
    "LOCKER DE CABECEIRA": {
        "component_type": "Locker de Cabeceira",
        "finishes": {
            2: ("Argila / Branco",           "madeirado"),
            3: ("Gianduia / Grafite / Preto","madeirado"),
            4: ("Madeirado",                 "madeirado"),
        },
    },
}

# SOFÁ NEO tratado separadamente por layout distinto.
SOFA_SHEET = "SOFÁ NEO"


# ---------------------------------------------------------------------------
# Utilitários
# ---------------------------------------------------------------------------

def norm(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ").strip())


def is_price(value: object) -> bool:
    if value is None:
        return False
    if isinstance(value, str) and ("#REF" in value or "tc<" in value.lower()):
        return False
    try:
        f = float(value)
        return 10.0 < f < 500_000.0
    except (TypeError, ValueError):
        return False


def extract_dimension(raw: str) -> str:
    """Extrai 'LxPxH' de strings como '1200 x 600 x 740mm'."""
    nums = re.findall(r"\d+", raw)
    if len(nums) >= 3:
        return f"{nums[0]}x{nums[1]}x{nums[2]}"
    if len(nums) == 2:
        return f"{nums[0]}x{nums[1]}"
    return raw.strip()


def extract_technical_description(rows: list[list[object]]) -> str | None:
    """Extrai o texto de descrição técnica do rodapé (linhas após 'DESCRIÇÃO TÉCNICA:')."""
    found = False
    parts: list[str] = []
    for row in rows:
        col0 = norm(row[0]) if row else ""
        if not found:
            if col0 == "DESCRIÇÃO TÉCNICA:":
                found = True
            continue
        # Após o marcador, coleta texto de col B (índice 1) ou col A (índice 0)
        col1 = norm(row[1]) if len(row) > 1 else ""
        text = col1 or col0
        if text:
            parts.append(text)
    return "\n".join(parts) if parts else None


def build_item(
    *,
    ref: str,
    component_type: str,
    description: str,
    dimension: str,
    finish: str,
    finish_group: str | None,
    price: float,
    confidence: float,
    notes: str | None,
    technical_description: str | None = None,
) -> dict[str, object]:
    item: dict[str, object] = {
        "ref": ref,
        "family": FAMILY,
        "product_context": None,
        "component_type": component_type,
        "description": description,
        "dimension": dimension,
        "finish": finish,
        "sku": None,
        "price": round(price, 2),
        "currency": "BRL",
        "confidence": confidence,
        "notes": notes,
        "technical_description": technical_description,
    }
    if finish_group is not None:
        item["finish_group"] = finish_group
    return item


# ---------------------------------------------------------------------------
# Parser de abas padrão (blocos de 4 linhas)
# ---------------------------------------------------------------------------

def parse_standard_sheet(
    rows: list[list[object]],
    sheet_name: str,
    config: dict,
    technical_description: str | None = None,
) -> list[dict[str, object]]:
    """Analisa abas com layout padrão de 4 linhas por item."""
    component_type = config["component_type"]
    finishes: dict[int, tuple[str, str]] = config["finishes"]
    finish_cols = list(finishes.keys())

    items: list[dict[str, object]] = []
    current_group: str = ""

    # Linhas de preço = linhas onde todas as colunas de acabamento têm preço válido.
    price_row_indices = [
        i for i, row in enumerate(rows)
        if len(row) > max(finish_cols)
        and all(is_price(row[c]) for c in finish_cols)
    ]

    for pi in price_row_indices:
        if pi < 3:
            continue  # precisa de pelo menos 3 linhas anteriores

        price_row = rows[pi]
        row_i   = rows[pi - 3]   # nome
        row_i1  = rows[pi - 2]   # dimensão
        row_i2  = rows[pi - 1]   # detalhe estrutura
        row_i3  = price_row      # detalhe config + preços

        # Atualiza grupo (MODELO) se col0 de qualquer linha do bloco tiver valor.
        for r in (row_i, row_i1, row_i2):
            g = norm(r[0]) if r else ""
            if g and g not in ("MODELO", "*imagens meramente ilustrativas.", "DESCRIÇÃO TÉCNICA:"):
                current_group = g

        name_raw   = norm(row_i[1])   if len(row_i)  > 1 else ""
        dim_raw    = norm(row_i1[1])  if len(row_i1) > 1 else ""
        detail1    = norm(row_i2[1])  if len(row_i2) > 1 else ""
        detail2    = norm(row_i3[1])  if len(row_i3) > 1 else ""

        # Reconstituir nome completo — às vezes col1 da linha de nome é idêntica
        # à linha anterior (repetição de cabeçalho de grupo). Usar como está.
        name = name_raw or current_group

        # Compor descrição natural com os detalhes disponíveis.
        parts = [p for p in (name, detail1, detail2) if p]
        description = " — ".join(parts) if len(parts) > 1 else (parts[0] if parts else name)

        dimension = extract_dimension(dim_raw) if dim_raw else ""

        ref = f"NOAR.xlsx!{sheet_name}!L{pi - 2},{pi + 1}"

        # Contexto extra para o revisor.
        ct_note = (
            f"Tipo de componente '{component_type}' é novo no catálogo "
            f"(família 'Linha Noar' ainda não existe)."
        )
        finish_note = (
            "Acabamento expresso como faixa de preço (ex. 'Argila / Branco' = "
            "mesmo preço para ambos os acabamentos); revisar mapeamento para "
            "acabamentos individuais do catálogo."
        )
        all_notes = f"{ct_note} {finish_note}"

        for col, (finish_name, finish_group) in finishes.items():
            price_val = float(price_row[col])
            items.append(
                build_item(
                    ref=ref,
                    component_type=component_type,
                    description=description,
                    dimension=dimension,
                    finish=finish_name,
                    finish_group=finish_group,
                    price=price_val,
                    confidence=0.82,
                    notes=all_notes,
                    technical_description=technical_description,
                )
            )

    return items


# ---------------------------------------------------------------------------
# Parser da aba SOFÁ NEO (layout especial)
# ---------------------------------------------------------------------------

def parse_sofa_sheet(rows: list[list[object]], technical_description: str | None = None) -> list[dict[str, object]]:
    """Analisa a aba SOFÁ NEO, que tem layout diferente das demais.

    Apenas a coluna de preço de 'LINHO / COURO ECO' (col2) possui valores
    válidos — as demais (col4, col5: COURO ECO, SPACER) contêm erros #REF!
    por referência quebrada a outras abas. Os 3 itens são extraídos apenas
    com a faixa de preço válida.

    Linhas 0-2 (título, aviso, cabeçalho) são ignoradas explicitamente para
    evitar que "DESCRIÇÃO" seja capturado como nome de produto.
    """
    items: list[dict[str, object]] = []

    # Tokens que não são nomes de produto: header, títulos de seção repetidos,
    # rótulos de marcador.
    # "DESCRIÇÃO TÉCNICA" aparece só no rodapé, após os preços, então não gera
    # itens. Os demais (título, cabeçalho) já são ignorados pelo guard i < 3.
    _SKIP_TOKENS = frozenset(["DESCRIÇÃO TÉCNICA"])

    valid_price_col = 2

    active_name: str = ""
    active_detail: str = ""

    for i, row in enumerate(rows):
        if i < 3:
            continue  # pula título, aviso e linha de cabeçalho

        col1 = norm(row[1]) if len(row) > 1 else ""
        col2_val = row[valid_price_col] if len(row) > valid_price_col else None

        if col1 and not is_price(col1) and col1.upper() not in _SKIP_TOKENS:
            if active_name:
                active_detail = col1
            else:
                active_name = col1

        if is_price(col2_val):
            price_val = float(col2_val)
            description = active_name if active_name else "Sofá NEO"
            if active_detail:
                description = f"{description} — {active_detail}"

            ref = f"NOAR.xlsx!{SOFA_SHEET}!L{i + 1}"
            note = (
                "Produto da linha Sofá NEO. "
                "Colunas de acabamento COURO ECO e SPACER contêm erros #REF! "
                "na planilha (referência quebrada a outras abas não incluídas nesta exportação); "
                "apenas o preço de 'LINHO / COURO ECO' foi extraído. "
                "Revisar disponibilidade de variantes de tecido/couro antes de publicar. "
                "Tipo de componente 'Sofá' e família 'Linha Noar' são novos no catálogo. "
                "finish_group não definido (acabamento em tecido/couro — não é madeirado nem metálico)."
            )
            items.append(
                build_item(
                    ref=ref,
                    component_type="Sofá",
                    description=description,
                    dimension="",
                    finish="Linho / Couro Eco",
                    finish_group=None,
                    price=price_val,
                    confidence=0.65,
                    notes=note,
                    technical_description=technical_description,
                )
            )
            active_name = ""
            active_detail = ""

    return items


# ---------------------------------------------------------------------------
# Construção do contrato
# ---------------------------------------------------------------------------

def build_contract() -> dict[str, object]:
    wb = load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    all_items: list[dict[str, object]] = []
    stats: list[str] = []

    for sheet_name, config in SHEET_CONFIGS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  [AVISO] Aba '{sheet_name}' não encontrada — ignorada.")
            continue
        ws = wb[sheet_name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        tech_desc = extract_technical_description(rows)
        items = parse_standard_sheet(rows, sheet_name, config, technical_description=tech_desc)
        print(f"  {sheet_name!r}: {len(items)} itens")
        stats.append(f"{sheet_name}: {len(items)}")
        all_items.extend(items)

    # SOFÁ NEO
    if SOFA_SHEET in wb.sheetnames:
        ws = wb[SOFA_SHEET]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        sofa_items = parse_sofa_sheet(rows, technical_description=extract_technical_description(rows))
        print(f"  {SOFA_SHEET!r}: {len(sofa_items)} itens")
        stats.append(f"{SOFA_SHEET}: {len(sofa_items)}")
        all_items.extend(sofa_items)

    all_items.sort(
        key=lambda it: (
            str(it.get("component_type", "")),
            str(it.get("dimension", "")),
            str(it.get("description", "")),
            str(it.get("finish", "")),
        )
    )

    return {
        "contract_version": "1.0",
        "source": {
            "description": SOURCE_PATH.name,
            "generated_by": "codex-parser-manual",
            "generated_at": (
                datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")
            ),
            "sheet_stats": stats,
        },
        "items": all_items,
    }


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main() -> None:
    print("Processando abas:")
    contract = build_contract()
    OUTPUT_PATH.write_text(
        json.dumps(contract, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    items = contract["items"]
    counts = Counter(str(it["component_type"]) for it in items)
    print(f"\nArquivo gerado: {OUTPUT_PATH}")
    print(f"Total de itens: {len(items)}")
    print("Por tipo de componente:")
    for ct, total in counts.most_common():
        print(f"  {ct}: {total}")


if __name__ == "__main__":
    main()
